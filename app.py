from flask import Flask, request, jsonify, render_template, redirect, url_for, session
from openpyxl import Workbook, load_workbook
from datetime import datetime
import os

app = Flask(__name__)
app.secret_key = "your_secret_key"  # needed for session login

file_name = "messages.xlsx"

# Create Excel file if it doesn't exist
if not os.path.exists(file_name):
    wb = Workbook()
    ws = wb.active
    ws.append(["Date", "Time", "Message"])
    wb.save(file_name)

# ---------- User Frontend ----------
@app.route('/')
def home():
    return render_template('index.html')

@app.route('/submit', methods=['POST'])
def submit():
    message = request.json.get('message')
    if not message:
        return jsonify({'status': 'error', 'message': 'Empty message'}), 400

    now = datetime.now()
    date_str = now.strftime("%Y-%m-%d")
    time_str = now.strftime("%H:%M:%S")

    wb = load_workbook(file_name)
    ws = wb.active
    ws.append([date_str, time_str, message])
    wb.save(file_name)

    return jsonify({'status': 'success', 'message': 'Message saved'})

@app.route('/messages', methods=['GET'])
def get_messages():
    wb = load_workbook(file_name)
    ws = wb.active
    messages = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        messages.append({'date': row[0], 'time': row[1], 'message': row[2]})
    return jsonify(messages)

# ---------- Admin Panel ----------
ADMIN_USERNAME = "admin"
ADMIN_PASSWORD = "password"  # change this to a strong password

@app.route('/admin', methods=['GET', 'POST'])
def admin():
    if request.method == 'POST':
        username = request.form.get('username')
        password = request.form.get('password')
        if username == ADMIN_USERNAME and password == ADMIN_PASSWORD:
            session['admin_logged_in'] = True
            return redirect(url_for('admin_dashboard'))
        else:
            return render_template('admin.html', error="Invalid credentials")
    return render_template('admin.html')

@app.route('/admin/dashboard')
def admin_dashboard():
    if not session.get('admin_logged_in'):
        return redirect(url_for('admin'))
    
    wb = load_workbook(file_name)
    ws = wb.active
    messages = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        messages.append({'date': row[0], 'time': row[1], 'message': row[2]})
    return render_template('admin.html', messages=messages, logged_in=True)

@app.route('/admin/logout')
def admin_logout():
    session.pop('admin_logged_in', None)
    return redirect(url_for('admin'))

import webbrowser

if __name__ == "__main__":
    # Open browser automatically
    webbrowser.open("http://127.0.0.1:5000/")
    app.run(debug=True)

